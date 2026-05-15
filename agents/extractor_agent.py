"""
Agent 2 — Requirements Extractor Agent

Takes the IngestedDocument and uses an LLM to extract all functional and
non-functional requirements. Outputs:
  • RequirementsAnalysis (in-memory)
  • Requirements Analysis .xlsx file (Stage 1 deliverable)
"""
from __future__ import annotations
import json
import pathlib
from typing import List

from models.schemas import IngestedDocument, Requirement, RequirementsAnalysis
from utils.llm import llm_json
import config


# ── System prompt ──────────────────────────────────────────────────────────────

_SYSTEM = """You are a senior IT architect specialising in requirements analysis.
Your job is to read RFP document sections and extract ALL functional and non-functional requirements.

For each requirement you must output a JSON object with these exact fields:
{
  "req_id":        "FA-N (functional) or NFR-N (non-functional)",
  "title":         "Short label (max 10 words)",
  "details":       "Full description including acceptance criteria",
  "feature_group": "The feature area / category this requirement belongs to",
  "note":          "HR Capability | IT capability | null",
  "questions":     "Any open clarification question, or null"
}

Rules:
- Number requirements sequentially: FA-1, FA-2, ... (continue from the last ID if given)
- Group requirements by the feature area they belong to (feature_group)
- Do NOT invent requirements — only extract what is explicitly stated in the text
- Return a JSON array of requirement objects. No other text.
"""


def _build_user_prompt(sections_text: str, start_id: int) -> str:
    return f"""Extract all requirements from the following RFP document sections.
Start numbering from FA-{start_id}.

=== RFP CONTENT ===
{sections_text}
=== END ===

Return a JSON array of requirement objects."""


# ── XLSX writer ───────────────────────────────────────────────────────────────

def _write_xlsx(requirements: List[Requirement], output_path: pathlib.Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header style
    header_fill = PatternFill("solid", fgColor="5B2D8E")  # purple
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "#", "Requirement", "Requirement details", "Assignment to feature",
        "Note", "solution", "comment", "Questions",
        "# interfaces", "#page", "#screen panel", "#ui int",
        "#business service", "data service", "#db table/view",
        "#db logic", "#workflow", "#report", "complexity"
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Data rows
    row_fill_even = PatternFill("solid", fgColor="F3E8FF")
    row_fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, req in enumerate(requirements, start=2):
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        data = [
            req.req_id, req.title, req.details, req.feature_group,
            req.note, req.solution, req.comment, req.questions,
            req.n_interfaces, req.n_pages, req.n_screen_panels, req.n_ui_int,
            req.n_biz_services, req.n_data_services, req.n_db_tables,
            req.n_db_logic, req.n_workflows, req.n_reports, req.complexity
        ]
        for col_idx, val in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    col_widths = [8, 25, 50, 30, 15, 25, 25, 25,
                  10, 8, 12, 8, 14, 12, 14, 10, 10, 8, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Application sheet (summary by feature)
    ws2 = wb.create_sheet("Application")
    ws2.append(["Assignment to feature", "Solution", "Notes"])
    feature_groups: dict[str, list] = {}
    for req in requirements:
        fg = req.feature_group
        if fg not in feature_groups:
            feature_groups[fg] = []
        if req.solution:
            feature_groups[fg].append(req.solution)
    for fg, sols in feature_groups.items():
        unique_sols = list(dict.fromkeys(sols))
        ws2.append([fg, ", ".join(unique_sols) if unique_sols else "", ""])

    # Assumptions sheet
    ws3 = wb.create_sheet("Assumptions")
    ws3.append(["Assumptions"])
    ws3.append(["To be filled during project kick-off based on finalised scope."])

    wb.save(str(output_path))
    print(f"  [Extractor] XLSX saved → {output_path.name}")


# ── Public interface ──────────────────────────────────────────────────────────

def run(doc: IngestedDocument) -> RequirementsAnalysis:
    """Extract requirements from all sections using LLM, chunk by chunk."""
    print(f"  [Extractor] Processing {len(doc.sections)} sections …")

    all_requirements: List[Requirement] = []
    req_counter = 1

    # Chunk sections to avoid token limits (max ~4000 chars per chunk)
    CHUNK_SIZE = 4000
    chunk_text = ""
    chunks: List[str] = []

    for section in doc.sections:
        section_text = f"\n## {section.title}\n{section.content}\n"
        if len(chunk_text) + len(section_text) > CHUNK_SIZE:
            if chunk_text:
                chunks.append(chunk_text)
            chunk_text = section_text
        else:
            chunk_text += section_text

    if chunk_text:
        chunks.append(chunk_text)

    print(f"  [Extractor] {len(chunks)} chunks to process …")

    for chunk_idx, chunk in enumerate(chunks):
        print(f"  [Extractor] LLM call — chunk {chunk_idx + 1}/{len(chunks)} …")
        try:
            raw_list = llm_json(
                system=_SYSTEM,
                user=_build_user_prompt(chunk, req_counter),
                max_tokens=4096,
            )
            if not isinstance(raw_list, list):
                print(f"  [Extractor] Warning: unexpected response type in chunk {chunk_idx+1}")
                continue

            for item in raw_list:
                req = Requirement(
                    req_id=item.get("req_id", f"FA-{req_counter}"),
                    title=item.get("title", ""),
                    details=item.get("details", ""),
                    feature_group=item.get("feature_group", "General"),
                    note=item.get("note"),
                    questions=item.get("questions"),
                )
                all_requirements.append(req)
                req_counter += 1

        except Exception as e:
            print(f"  [Extractor] Error in chunk {chunk_idx+1}: {e}")
            continue

    # Re-number sequentially to avoid duplicates from chunking
    seen_ids: set[str] = set()
    fa_counter = 1
    for req in all_requirements:
        if req.req_id in seen_ids:
            req.req_id = f"FA-{fa_counter}"
        seen_ids.add(req.req_id)
        fa_counter += 1

    print(f"  [Extractor] → {len(all_requirements)} requirements extracted")

    analysis = RequirementsAnalysis(requirements=all_requirements)

    # Write Stage 1 XLSX output
    xlsx_path = config.OUTPUTS_DIR / "Requirements_Analysis.xlsx"
    _write_xlsx(all_requirements, xlsx_path)

    return analysis
