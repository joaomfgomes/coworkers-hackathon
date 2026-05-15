"""
Agent 3 — Solution Mapper Agent

For each requirement, retrieves the best-matching solution from the
knowledge base and uses the LLM to finalise the mapping.

Outputs: RequirementsAnalysis with solution/comment/complexity columns filled.
Also updates the XLSX with the enriched data.
"""
from __future__ import annotations
import json
from typing import List

from models.schemas import Requirement, RequirementsAnalysis
from knowledge_base.kb import retrieve_top, feature_group_lookup, get_all
from utils.llm import llm_json
import config


_SYSTEM = """You are a senior SAP/cloud architect mapping HR RFP requirements to technology solutions.

Given a requirement and a list of candidate solutions from our knowledge base,
select the BEST matching solution and return:
{
  "solution":   "exact solution name from candidates",
  "comment":    "RICEFW to be considered | null",
  "complexity": "Low | Medium | High",
  "n_interfaces": integer or null,
  "n_pages": integer or null,
  "n_screen_panels": integer or null,
  "n_ui_int": integer or null,
  "n_biz_services": integer or null,
  "n_data_services": integer or null,
  "n_db_tables": integer or null,
  "n_db_logic": integer or null,
  "n_workflows": integer or null,
  "n_reports": integer or null
}

Rules:
- Prefer the candidate whose feature_group most closely matches
- If no candidate fits, propose "Custom Development (BTP)"
- Return ONLY the JSON object, no other text
"""


def _map_one(req: Requirement) -> None:
    """Enrich a single requirement in-place."""
    # First try exact feature group match
    kb_entry = feature_group_lookup(req.feature_group)

    if kb_entry:
        # Exact match — use directly without extra LLM call
        req.solution = kb_entry["solution"]
        req.comment  = kb_entry.get("comment")
        req.complexity = _infer_complexity(req)
        return

    # No exact match — retrieve top candidates and ask LLM
    candidates = retrieve_top(f"{req.feature_group} {req.details}", top_k=3)
    candidates_text = json.dumps(
        [{"feature_group": c["feature_group"], "solution": c["solution"]} for c in candidates],
        indent=2
    )

    user_prompt = f"""Requirement:
ID: {req.req_id}
Title: {req.title}
Feature group: {req.feature_group}
Details: {req.details[:800]}

Candidate solutions from knowledge base:
{candidates_text}

Select the best solution and return the JSON mapping."""

    try:
        result = llm_json(_SYSTEM, user_prompt, max_tokens=512)
        req.solution        = result.get("solution", candidates[0]["solution"] if candidates else "Custom Development (BTP)")
        req.comment         = result.get("comment")
        req.complexity      = result.get("complexity")
        req.n_interfaces    = result.get("n_interfaces")
        req.n_pages         = result.get("n_pages")
        req.n_screen_panels = result.get("n_screen_panels")
        req.n_ui_int        = result.get("n_ui_int")
        req.n_biz_services  = result.get("n_biz_services")
        req.n_data_services = result.get("n_data_services")
        req.n_db_tables     = result.get("n_db_tables")
        req.n_db_logic      = result.get("n_db_logic")
        req.n_workflows     = result.get("n_workflows")
        req.n_reports       = result.get("n_reports")
    except Exception as e:
        print(f"  [Mapper] Warning: LLM error for {req.req_id}: {e}")
        req.solution   = candidates[0]["solution"] if candidates else "Custom Development (BTP)"
        req.complexity = "Medium"


def _infer_complexity(req: Requirement) -> str:
    """Simple heuristic complexity based on details length and keywords."""
    text = req.details.lower()
    if any(w in text for w in ["complex", "multiple", "integration", "migration", "custom"]):
        return "High"
    if any(w in text for w in ["simple", "standard", "template", "existing"]):
        return "Low"
    return "Medium"


def _update_xlsx(requirements: List[Requirement]) -> None:
    """Update the XLSX with solution/comment/complexity columns."""
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment

    xlsx_path = config.OUTPUTS_DIR / "Requirements_Analysis.xlsx"
    if not xlsx_path.exists():
        print("  [Mapper] XLSX not found — skipping update")
        return

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    # Column indices (1-based): solution=6, comment=7, complexity=19
    # interfaces=9, page=10, screen=11, ui_int=12, biz=13, data=14, db_tbl=15, db_logic=16, wf=17, report=18
    col_map = {
        "solution": 6, "comment": 7, "questions": 8,
        "n_interfaces": 9, "n_pages": 10, "n_screen_panels": 11, "n_ui_int": 12,
        "n_biz_services": 13, "n_data_services": 14, "n_db_tables": 15,
        "n_db_logic": 16, "n_workflows": 17, "n_reports": 18, "complexity": 19,
    }

    req_by_id = {r.req_id: r for r in requirements}

    for row_idx in range(2, ws.max_row + 1):
        req_id_cell = ws.cell(row=row_idx, column=1).value
        if req_id_cell and req_id_cell in req_by_id:
            req = req_by_id[req_id_cell]
            for field, col in col_map.items():
                val = getattr(req, field, None)
                if val is not None:
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Also refresh Application sheet
    if "Application" in wb.sheetnames:
        ws_app = wb["Application"]
        # Clear and rewrite
        for row in ws_app.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        feature_map: dict[str, str] = {}
        for req in requirements:
            if req.solution:
                feature_map[req.feature_group] = req.solution
        for r_idx, (fg, sol) in enumerate(feature_map.items(), start=2):
            ws_app.cell(row=r_idx, column=1, value=fg)
            ws_app.cell(row=r_idx, column=2, value=sol)

    wb.save(str(xlsx_path))
    print(f"  [Mapper] XLSX enriched with solutions → {xlsx_path.name}")


# ── Public interface ──────────────────────────────────────────────────────────

def run(analysis: RequirementsAnalysis) -> RequirementsAnalysis:
    """Map each requirement to a solution using KB + LLM."""
    reqs = analysis.requirements
    print(f"  [Mapper] Mapping solutions for {len(reqs)} requirements …")

    # Group by feature_group to batch LLM calls
    feature_groups: dict[str, list] = {}
    for req in reqs:
        feature_groups.setdefault(req.feature_group, []).append(req)

    total = len(reqs)
    done = 0
    for fg, group in feature_groups.items():
        # Try exact KB match first (no LLM needed for known groups)
        kb_entry = feature_group_lookup(fg)
        if kb_entry:
            for req in group:
                req.solution   = kb_entry["solution"]
                req.comment    = kb_entry.get("comment")
                req.complexity = _infer_complexity(req)
                done += 1
        else:
            # Use LLM for unknown feature groups
            for req in group:
                _map_one(req)
                done += 1
        print(f"  [Mapper] {done}/{total} mapped …")

    _update_xlsx(reqs, )

    return RequirementsAnalysis(requirements=reqs)
